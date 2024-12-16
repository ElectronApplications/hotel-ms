from rest_framework.viewsets import GenericViewSet

from django.db.models import Avg, Count, Max, Min
from django.http import HttpResponse

from tempfile import NamedTemporaryFile
from openpyxl import Workbook

from hotel.models.accomodation import Accomodation
from hotel.permissions import IsAdmin

class StatsViewset(GenericViewSet):
    queryset = Accomodation.objects.all()
    permission_classes = [IsAdmin]

    def list(self, request, *args, **kwargs):
        accomodationPriceStats = self.get_queryset().aggregate(
            count=Count("*"),
            avg=Avg("price_to_pay"),
            max=Max("price_to_pay"),
            min=Min("price_to_pay"),
        )

        wb = Workbook()
        ws = wb.active

        ws["A1"] = "Count"
        ws["A2"] = "Average price"
        ws["A3"] = "Max price"
        ws["A4"] = "Min price"

        ws["B1"] = accomodationPriceStats["count"]
        ws["B2"] = accomodationPriceStats["avg"]
        ws["B3"] = accomodationPriceStats["max"]
        ws["B4"] = accomodationPriceStats["min"]

        stream: bytes
        with NamedTemporaryFile() as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()

        response = HttpResponse(content=stream, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = "attachment; filename=stats.xlsx"
        return response
